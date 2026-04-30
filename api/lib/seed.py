"""
nim-agents-ops api/lib/seed.py

loads ds_routing (156 UAE rows) + vendor_routing (87 vendors) from the matrix
xlsx file under pinned/. idempotent — wipes and re-inserts each table.

usage:
  python -m api.lib.seed
"""
import os, sys
import openpyxl

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", ".."))
MATRIX_XLSX = os.path.join(ROOT, "pinned", "nim-agents-ops-ds-matrix-v09-2026-04-27.xlsx")
VENDOR_XLSX = os.path.join(ROOT, "pinned", "vendor_directory_for_validation.xlsx")

sys.path.insert(0, ROOT)
from api.lib.db import init_db, conn


def seed_ds_routing():
    wb = openpyxl.load_workbook(MATRIX_XLSX, data_only=True)
    ws = wb["ds_area_manager"]
    rows = list(ws.iter_rows(values_only=True))
    header = None
    data = []
    for r in rows:
        if header is None:
            if r and r[0] == "geo":
                header = list(r)
            continue
        if r and r[0] in ("uae", "ksa", "egy"):
            data.append(r)

    with conn() as c:
        c.execute("DELETE FROM ds_routing")
        for r in data:
            d = dict(zip(header, r))
            c.execute(
                """INSERT INTO ds_routing
                (ds_code, ds_name, geo, city, ds_status, am_name, am_email,
                 asst_mgr, supervisor, tl_name)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (
                    d.get("ds_code"), d.get("ds_name"), d.get("geo"),
                    d.get("city"), d.get("ds_status"),
                    d.get("area_manager_name"), d.get("area_manager_email"),
                    d.get("asst_manager"), d.get("supervisor"), d.get("tl_name"),
                )
            )
    return len(data)


def seed_vendor_routing():
    wb = openpyxl.load_workbook(VENDOR_XLSX, data_only=True)
    ws = wb["vendors"]
    rows = list(ws.iter_rows(values_only=True))
    header = None
    data = []
    for r in rows:
        if header is None:
            if r and r[0] == "#":
                header = list(r)
            continue
        if r and r[1] is not None:
            data.append(r)

    with conn() as c:
        c.execute("DELETE FROM vendor_routing")
        count = 0
        for r in data:
            d = dict(zip(header, r))
            id_vendor = d.get("id_vendor")
            if id_vendor is None:
                continue
            try:
                id_vendor = int(id_vendor)
            except (TypeError, ValueError):
                continue
            email_field = None
            for k in d.keys():
                if k and "email" in str(k).lower():
                    email_field = k
                    break
            email = d.get(email_field) if email_field else None
            in_scope = 0 if id_vendor == 143 else 1
            c.execute(
                """INSERT OR REPLACE INTO vendor_routing
                (id_vendor, shortcode, vendor_name, vendor_email, vendor_type, in_scope)
                VALUES (?,?,?,?,?,?)""",
                (id_vendor, d.get("shortcode"), d.get("vendor_name"),
                 email, d.get("vendor_type"), in_scope)
            )
            count += 1
    return count


def main():
    init_db()
    n_ds = seed_ds_routing()
    n_v = seed_vendor_routing()
    print(f"seeded {n_ds} ds_routing rows + {n_v} vendor_routing rows "
          f"(id_vendor=143 marked out_of_scope)")


if __name__ == "__main__":
    main()
