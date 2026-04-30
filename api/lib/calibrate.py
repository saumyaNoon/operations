"""
nim-agents-ops api/lib/calibrate.py

L7D distribution per agent + proposed thresholds. usage:
  python -m api.lib.calibrate

writes outputs/Morpheus/morpheus_thresholds_calibration.xlsx
"""
import os, sys
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", ".."))
sys.path.insert(0, ROOT)

from api.lib.bigquery_client import run as bq_run

OUT_DIR = r"C:\Users\vnagar\Documents\Claude\outputs\Morpheus"
OUT_PATH = os.path.join(OUT_DIR, "morpheus_thresholds_calibration.xlsx")


def _pcts(values, pcts=(50, 75, 90, 95, 99)):
    if not values:
        return {f"p{p}": None for p in pcts}
    s = sorted(values)
    n = len(s)
    out = {}
    for p in pcts:
        idx = max(0, min(n - 1, int(round(p / 100 * (n - 1)))))
        out[f"p{p}"] = s[idx]
    return out


# ──────────────────────────────────────────────────────────────────────────────
# distribution pulls per agent (L7D)
# ──────────────────────────────────────────────────────────────────────────────
def dist_attendance_cc(d_end):
    sql = f"""
    WITH manpower AS (
      SELECT b.wh_code, b.created_date,
        COUNT(DISTINCT b.employee_id) AS active,
        COUNT(DISTINCT CASE WHEN b.punch_status = 'Week Off' THEN b.employee_id END) AS week_off,
        COUNT(DISTINCT CASE WHEN b.punch_status = 'Annual Leave' THEN b.employee_id END) AS al_leave,
        COUNT(DISTINCT CASE WHEN LOWER(b.punch_status) LIKE '%leave' AND b.punch_status <> 'Annual Leave' THEN b.employee_id END) AS other_leave,
        COUNT(DISTINCT CASE WHEN b.punch_status NOT IN ('Punched Properly','Week Off','Annual Leave') AND LOWER(b.punch_status) NOT LIKE '%leave' THEN b.employee_id END) AS absent_other
      FROM `noonbinimksa.Stores.Biometric_base_v2_3` b
      LEFT JOIN `noonbinimksa.Stores.warehouse` w ON b.wh_code = w.partner_wh_code
      WHERE b.created_date BETWEEN DATE_SUB(DATE('{d_end}'), INTERVAL 7 DAY) AND DATE('{d_end}')
        AND LOWER(w.country_code) = 'ae'
        AND UPPER(b.HR_Desig) NOT LIKE '%TEMP%' AND b.HR_Desig IS NOT NULL
      GROUP BY b.wh_code, b.created_date
    )
    SELECT ROUND(SAFE_DIVIDE(absent_other, NULLIF(active - week_off - al_leave - other_leave, 0)) * 100, 2) AS absent_pct
    FROM manpower
    WHERE (active - week_off - al_leave - other_leave) > 0
    """
    return bq_run(sql)


def dist_iph_pickers(d_end):
    sql = f"""
    SELECT
      store AS ds_code, date,
      SAFE_DIVIDE(SUM(outbound_qty) * 3600.0, NULLIF(SUM(outbound_time_picking), 0)) AS iph
    FROM `noonbinimksa.darkstore.ipp_daily_ae`
    WHERE date BETWEEN DATE_SUB(DATE('{d_end}'), INTERVAL 7 DAY) AND DATE('{d_end}')
    GROUP BY store, date
    HAVING SUM(outbound_time_picking) > 0
    """
    return bq_run(sql)


def dist_iph_putaway(d_end):
    sql = f"""
    SELECT
      dist_partner_code AS ds_code, DATE(date) AS date,
      SAFE_DIVIDE(SUM(completed_count) * 3600.0, NULLIF(SUM(total_sec), 0)) AS iph
    FROM `noonbinimops.fulfillment.ipp`
    WHERE DATE(date) BETWEEN DATE_SUB(DATE('{d_end}'), INTERVAL 7 DAY) AND DATE('{d_end}')
    GROUP BY dist_partner_code, DATE(date)
    HAVING SUM(total_sec) > 0
    """
    return bq_run(sql)


def dist_skips_picker(d_end):
    sql = f"""
    WITH s AS (
      SELECT DATE(date_) AS d, partner_wh_code AS ds_code, SUM(items) AS skips
      FROM `noonbinimksa.darkstore.daily_manual_skips_hourly_uae_1`
      WHERE DATE(date_) BETWEEN DATE_SUB(DATE('{d_end}'), INTERVAL 7 DAY) AND DATE('{d_end}')
      GROUP BY d, partner_wh_code
    ),
    p AS (
      SELECT date AS d, store AS ds_code, SUM(outbound_qty) AS picked
      FROM `noonbinimksa.darkstore.ipp_daily_ae`
      WHERE date BETWEEN DATE_SUB(DATE('{d_end}'), INTERVAL 7 DAY) AND DATE('{d_end}')
      GROUP BY d, store
    )
    SELECT ROUND(SAFE_DIVIDE(s.skips, NULLIF(p.picked, 0)) * 100, 4) AS skip_pct, s.skips
    FROM s LEFT JOIN p USING (d, ds_code)
    WHERE p.picked > 0 AND s.skips > 0
    """
    return bq_run(sql)


def dist_defects(d_end):
    """ds-level overall defect_rate_pct per ds × day (matches new agent_05)."""
    sql = f"""
    WITH d AS (
      SELECT complain_date, partner_wh_code, COUNT(DISTINCT order_nr) AS def_orders
      FROM `noonbinimksa.darkstore.complains_raw_all`
      WHERE complain_date BETWEEN DATE_SUB(DATE('{d_end}'), INTERVAL 7 DAY) AND DATE('{d_end}')
        AND country_code = 'ae'
      GROUP BY complain_date, partner_wh_code
    ),
    o AS (
      SELECT date AS d, store, SUM(total_orders) AS orders
      FROM `noonbinimksa.darkstore.ipp_daily_ae`
      WHERE date BETWEEN DATE_SUB(DATE('{d_end}'), INTERVAL 7 DAY) AND DATE('{d_end}')
      GROUP BY d, store
    )
    SELECT ROUND(SAFE_DIVIDE(d.def_orders, NULLIF(o.orders, 0)) * 100, 4) AS defect_rate_pct
    FROM d JOIN o ON d.complain_date = o.d AND d.partner_wh_code = o.store
    WHERE o.orders > 0
    """
    return bq_run(sql)


def dist_fefo(d_end):
    """fefo % of orders per ds × day (matches new agent_06)."""
    sql = f"""
    WITH logs AS (
      SELECT DISTINCT a.* FROM `noonbinimdwh.modelling.fifo_report_logs` a
      JOIN (
        SELECT wh_code, sku, date_, MAX(updated_at) AS updated_at
        FROM `noonbinimdwh.modelling.fifo_report_logs`
        WHERE date_ BETWEEN DATE_SUB(DATE('{d_end}'), INTERVAL 7 DAY) AND DATE('{d_end}')
        GROUP BY wh_code, sku, date_
      ) b USING (wh_code, sku, date_, updated_at)
      WHERE a.country_code = 'ae'
    ),
    fefo AS (
      SELECT ds_code, date_,
             SUM(CASE WHEN fefo_breach THEN 1 ELSE 0 END) AS skus_breached
      FROM logs GROUP BY ds_code, date_ HAVING skus_breached > 0
    ),
    gmv AS (
      SELECT wb.partner_wh_code AS ds_code, g.created_date_uae AS date_,
             SUM(g.order_nr_cnt) AS orders
      FROM `noonbinimksa.darkstore.odr_gmv_uae` g
      LEFT JOIN `noonbinimdwh.chatbot.warehouse_base_table` wb ON wb.wh_code = g.wh_code
      WHERE g.created_date_uae BETWEEN DATE_SUB(DATE('{d_end}'), INTERVAL 7 DAY) AND DATE('{d_end}')
      GROUP BY wb.partner_wh_code, g.created_date_uae
    )
    SELECT ROUND(SAFE_DIVIDE(f.skus_breached, NULLIF(g.orders, 0))*100, 4) AS fefo_pct_orders
    FROM fefo f LEFT JOIN gmv g USING (ds_code, date_)
    WHERE g.orders > 0
    """
    return bq_run(sql)


def dist_adjustments(d_end):
    """adj_pct = adj_value / live_inv_value, using cost_price_retail."""
    sql = f"""
    WITH adj AS (
      SELECT put_date, wh_code,
        ROUND(SUM(IFNULL(CAST(positive_variance_value AS FLOAT64), 0))
             + SUM(IFNULL(CAST(negative_variance_value AS FLOAT64), 0)), 2) AS adj_value
      FROM `noonbinimops.fulfillment.adjustments_master_uae`
      WHERE put_date BETWEEN DATE_SUB(DATE('{d_end}'), INTERVAL 7 DAY) AND DATE('{d_end}')
      GROUP BY put_date, wh_code HAVING adj_value > 0
    ),
    cost AS (
      SELECT sku, ANY_VALUE(cost_price) AS cost_price
      FROM (SELECT sku, cost_price, ROW_NUMBER() OVER (PARTITION BY sku ORDER BY year_month DESC) AS rn
            FROM `noonbinimprc.pricing.cost_price_retail` WHERE country_code = 'ae' AND cost_price > 0)
      WHERE rn = 1 GROUP BY sku
    ),
    inv AS (
      SELECT q.warehouse AS wh_code,
             ROUND(SUM(q.qty * IFNULL(CAST(c.cost_price AS FLOAT64), 0)), 2) AS live_inv_value
      FROM `noonbinimksa.Stores.wh_loc_qty` q
      LEFT JOIN cost c ON c.sku = q.zsku
      WHERE LOWER(q.country_code) = 'ae' GROUP BY q.warehouse
    )
    SELECT ROUND(SAFE_DIVIDE(adj.adj_value, NULLIF(inv.live_inv_value, 0)) * 100, 4) AS adj_pct
    FROM adj LEFT JOIN inv USING (wh_code) WHERE inv.live_inv_value > 0
    """
    return bq_run(sql)


def dist_putaway_delays(d_end):
    """breach_qty per (ds × storage_type) on the live snapshot."""
    sql = f"""
    WITH typed AS (
      SELECT ds_code, ageing_in_mins, items_pending,
        CASE
          WHEN src_wh_code IN ('AUHFKZ01','AUHFKZ02','AUHFKZ03') THEN 'frozen'
          WHEN src_wh_code IN ('DXBID01','DXBID02','AUHID01','AUHID02','DXSID01','DXSID02') THEN 'ambient'
          WHEN src_wh_code IN ('AUHID03','AUHID04','DXBID03','DXBID04') THEN 'chiller'
          WHEN src_wh_code IN ('AUHID07','DXBID07') THEN 'ultrafresh'
          WHEN src_wh_code IN ('DXBFNV01') THEN 'fnv'
          ELSE 'others'
        END AS storage_type
      FROM `noonbinimops.fulfillment.putaway_pendency_v2`
      WHERE date BETWEEN DATE_SUB(CURRENT_DATE(), INTERVAL 7 DAY) AND CURRENT_DATE()
        AND LOWER(country_code) = 'ae'
    )
    SELECT
      SUM(CASE
            WHEN storage_type = 'ambient' AND ageing_in_mins > 360 THEN items_pending
            WHEN storage_type = 'chiller' AND ageing_in_mins > 60  THEN items_pending
            WHEN storage_type IN ('frozen','ultrafresh','fnv') AND ageing_in_mins > 30 THEN items_pending
            ELSE 0 END) AS breach_qty
    FROM typed
    WHERE storage_type <> 'others'
    GROUP BY ds_code, storage_type
    HAVING breach_qty > 0
    """
    return bq_run(sql)


def dist_missing_inv(d_end):
    """missing_value as % of GMV per ds × date."""
    sql = f"""
    WITH s AS (
      SELECT created_date, partner_warehouse_code AS ds_code, psku_code,
             SUM(COALESCE(variance_, 0)) AS net_variance
      FROM `noonbinimksa.darkstore.stock_take_base`
      WHERE created_date BETWEEN DATE_SUB(DATE('{d_end}'), INTERVAL 7 DAY) AND DATE('{d_end}')
        AND LOWER(country_code) = 'ae'
      GROUP BY created_date, partner_warehouse_code, psku_code
    ),
    psku_map AS (
      SELECT psku_code, ANY_VALUE(zsku_child) AS zsku FROM `noondwh.zsku_catexsp.psku` GROUP BY psku_code
    ),
    cost AS (
      SELECT sku, ANY_VALUE(cost_price) AS cost_price
      FROM (SELECT sku, cost_price, ROW_NUMBER() OVER (PARTITION BY sku ORDER BY year_month DESC) AS rn
            FROM `noonbinimprc.pricing.cost_price_retail` WHERE country_code = 'ae' AND cost_price > 0)
      WHERE rn = 1 GROUP BY sku
    ),
    valued AS (
      SELECT s.created_date, s.ds_code,
             SUM(ABS(s.net_variance) * IFNULL(CAST(c.cost_price AS FLOAT64),0)) AS missing_value
      FROM s LEFT JOIN psku_map pm ON pm.psku_code = s.psku_code
      LEFT JOIN cost c ON c.sku = pm.zsku
      GROUP BY s.created_date, s.ds_code
    ),
    gmv AS (
      SELECT g.created_date_uae AS d, wb.partner_wh_code AS ds_code, SUM(g.gmv) AS store_gmv
      FROM `noonbinimksa.darkstore.odr_gmv_uae` g
      LEFT JOIN `noonbinimdwh.chatbot.warehouse_base_table` wb ON wb.wh_code = g.wh_code
      WHERE g.created_date_uae BETWEEN DATE_SUB(DATE('{d_end}'), INTERVAL 7 DAY) AND DATE('{d_end}')
      GROUP BY g.created_date_uae, wb.partner_wh_code
    )
    SELECT ROUND(SAFE_DIVIDE(v.missing_value, NULLIF(g.store_gmv, 0)) * 100, 4) AS missing_value_pct
    FROM valued v LEFT JOIN gmv g ON v.created_date = g.d AND v.ds_code = g.ds_code
    WHERE g.store_gmv > 0 AND v.missing_value > 0
    """
    return bq_run(sql)


def dist_audit(d_end):
    sql = f"""
    WITH latest AS (
      SELECT *, MAX(audit_date) OVER () AS max_d
      FROM `noonbinimksa.darkstore.historic_score1`
      WHERE country_code = 'AE'
    )
    SELECT score1 FROM latest WHERE audit_date = max_d AND score1 IS NOT NULL
    """
    return bq_run(sql)


def dist_stocktake_adherence(d_end):
    sql = f"""
    SELECT
      w.partner_warehouse_code AS ds_code, DATE(j.created_at) AS d,
      ROUND(SAFE_DIVIDE(SUM(CASE WHEN j.id_status = 3 THEN 1 ELSE 0 END), COUNT(*)) * 100, 1) AS adherence_pct,
      COUNT(*) AS jobs
    FROM `noondwh.mxdcss_dcss.job` j
    JOIN `noondwh.mxdcss_dcss.warehouse` w ON w.id_warehouse = j.id_warehouse
    WHERE DATE(j.created_at) BETWEEN DATE_SUB(DATE('{d_end}'), INTERVAL 7 DAY) AND DATE('{d_end}')
      AND w.id_country = 1 AND j.id_job_subtype = 10
    GROUP BY w.partner_warehouse_code, DATE(j.created_at)
    HAVING jobs >= 5
    """
    rows = bq_run(sql)
    # gap = 100 - adherence
    return [{**r, "adherence_gap_pct": 100 - (r["adherence_pct"] or 0)} for r in rows]


# ──────────────────────────────────────────────────────────────────────────────
# config
# ──────────────────────────────────────────────────────────────────────────────
SPECS = [
    {"agent": "agent_01_attendance", "metric": "absent_pct (cc-grain ds × day)",
     "dist_fn": dist_attendance_cc, "field": "absent_pct",
     "current_t3": "absent% > 8% OR count > 10", "current_t2": "> 5% OR > 5", "current_t1": "> 3%"},
    {"agent": "agent_02_iph_pickers", "metric": "iph (ds × date, picker outbound)",
     "dist_fn": dist_iph_pickers, "field": "iph",
     "current_t3": "iph < bucket.p10 (lower=worse)", "current_t2": "p10–p20", "current_t1": "p20–p50"},
    {"agent": "agent_03_iph_putaway", "metric": "iph (ds × date, putaway inbound)",
     "dist_fn": dist_iph_putaway, "field": "iph",
     "current_t3": "iph < bucket.p10", "current_t2": "p10–p20", "current_t1": "p20–p50"},
    {"agent": "agent_04_skips_picker", "metric": "skip_pct (ds × date)",
     "dist_fn": dist_skips_picker, "field": "skip_pct",
     "current_t3": "> 0.20% AND ≥2", "current_t2": "0.15–0.20% AND ≥2", "current_t1": "0.10–0.15% AND ≥2"},
    {"agent": "agent_05_defects", "metric": "defect_rate_pct (ds × date)",
     "dist_fn": dist_defects, "field": "defect_rate_pct",
     "current_t3": "> 0.80%", "current_t2": "0.60–0.80%", "current_t1": "0.50–0.60%"},
    {"agent": "agent_06_fefo", "metric": "fefo skus_breached / orders × 100",
     "dist_fn": dist_fefo, "field": "fefo_pct_orders",
     "current_t3": "> 1.00%", "current_t2": "0.50–1.00%", "current_t1": "0.30–0.50%"},
    {"agent": "agent_07_adjustments", "metric": "adj_value / live_inv_value × 100",
     "dist_fn": dist_adjustments, "field": "adj_pct",
     "current_t3": "> 0.50% AND > $2000", "current_t2": "0.25–0.50% AND > $1000", "current_t1": "> 0.10% AND > $500"},
    {"agent": "agent_08_putaway_delays", "metric": "breach_qty (ds × storage_type)",
     "dist_fn": dist_putaway_delays, "field": "breach_qty",
     "current_t3": "> 50", "current_t2": "25–50", "current_t1": "10–25"},
    {"agent": "agent_09_missing_inventory", "metric": "missing_value / store_gmv × 100",
     "dist_fn": dist_missing_inv, "field": "missing_value_pct",
     "current_t3": "> 0.30%", "current_t2": "0.20–0.30%", "current_t1": "0.10–0.20%"},
    {"agent": "agent_10_skips_stocktake", "metric": "100 - jobs_adherence_pct (ds × day)",
     "dist_fn": dist_stocktake_adherence, "field": "adherence_gap_pct",
     "current_t3": "adherence < 50% (gap > 50)", "current_t2": "50–70% (gap 30–50)", "current_t1": "70–85% (gap 15–30)"},
    {"agent": "agent_11_audit_scores", "metric": "score1 (latest weekly audit per ds)",
     "dist_fn": dist_audit, "field": "score1",
     "current_t3": "score1 < 0.85 OR fails_4w ≥ 3",
     "current_t2": "score1 < 0.90 OR fails_4w == 2",
     "current_t1": "score1 < 0.95 OR fails_4w == 1"},
]


def _suggest(pcts, agent_id):
    if agent_id in ("agent_02_iph_pickers", "agent_03_iph_putaway"):
        # iph: lower=worse, suggest already wired (p10/p20/p50)
        return ("p10 (in-bucket)", "p20 (in-bucket)", "p50 (in-bucket)")
    if agent_id == "agent_11_audit_scores":
        return ("< 0.85", "0.85–0.90", "0.90–0.95")
    # higher = worse: t3 = p90, t2 = p75, t1 = p50
    return pcts.get("p90"), pcts.get("p75"), pcts.get("p50")


def main():
    d_end = str(date.today() - timedelta(days=1))
    print(f"calibrating against L7D ending {d_end} …")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "thresholds"

    bold = Font(bold=True, color="FFFFFF", name="DM Sans", size=11)
    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    border = Border(left=Side(style="thin", color="E5E7EB"),
                    right=Side(style="thin", color="E5E7EB"),
                    top=Side(style="thin", color="E5E7EB"),
                    bottom=Side(style="thin", color="E5E7EB"))

    headers = [
        "agent", "metric", "rows (L7D)",
        "p50", "p75", "p90", "p95", "p99",
        "current t3", "current t2", "current t1",
        "L7D-suggested t3", "L7D-suggested t2", "L7D-suggested t1",
        "vardan decision (t3)", "vardan decision (t2)", "vardan decision (t1)",
        "notes",
    ]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = bold; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    row = 2
    for spec in SPECS:
        try:
            data = spec["dist_fn"](d_end)
        except Exception as e:
            ws.cell(row=row, column=1, value=spec["agent"])
            ws.cell(row=row, column=2, value=spec["metric"])
            ws.cell(row=row, column=18, value=f"ERR: {str(e)[:200]}")
            row += 1; continue

        values = [r.get(spec["field"]) for r in data if r.get(spec["field"]) is not None]
        pcts = _pcts(values, (5, 10, 20, 50, 75, 90, 95, 99))
        suggested = _suggest(pcts, spec["agent"])

        cells = [
            spec["agent"], spec["metric"], len(values),
            pcts.get("p50"), pcts.get("p75"), pcts.get("p90"),
            pcts.get("p95"), pcts.get("p99"),
            spec["current_t3"], spec["current_t2"], spec["current_t1"],
            suggested[0], suggested[1], suggested[2],
            "", "", "",
            "",
        ]
        for c, v in enumerate(cells, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = border
            cell.font = Font(name="DM Sans", size=10)
            if isinstance(v, float):
                cell.number_format = "0.000"
        row += 1

    widths = {1: 28, 2: 42, 3: 12, 4: 10, 5: 10, 6: 10, 7: 10, 8: 10,
              9: 32, 10: 32, 11: 32, 12: 18, 13: 18, 14: 18, 15: 22, 16: 22, 17: 22, 18: 50}
    for col, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "B2"

    os.makedirs(OUT_DIR, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
